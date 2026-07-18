#!/usr/bin/env python3
"""Run eval.py for every dataset/method and export the metrics to Excel."""

import argparse
import json
import re
import subprocess
import sys
import math
import statistics
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import sample_filter


PROJECT_DIR = Path(__file__).resolve().parent
TASK_SIZES = {
    "GSM8K": 500,
    "SVAMP": 500,
    "AQuA": 254,
    "MultiArith": 500,
    "AddSub": 395,
    "SingleEq": 500,
    "ARC-c": 500,
    "StrategyQA": 500,
    "Colored_Objects": 250,
    "Penguins": 146,
}
METHODS = [
    ("single_agent", "Zero-shot CoT"),
    ("self_correction", "Self-correct"),
    ("majority", "Multi-agent Majority"),
    ("debate", "Multi-agent Debate"),
    ("peer_review", "Ours"),
    ("feedback", "Ours (w/o confidence)"),
    ("ablation_solution", "Ours (w/o solution)"),
]
PAPER_ROWS = [
    ("Zero-shot CoT", "single_agent"),
    ("Self-correct", "self_correction"),
    ("Multi-agent Majority", "majority"),
    ("Multi-agent Debate", "debate"),
    ("Ours", "peer_review"),
    ("Ours (w/o confidence)", "feedback"),
    ("Ours (w/o solution)", "ablation_solution"),
]
FILE_METHOD = {"majority": "single_agent"}
METRIC_RE = re.compile(
    r"(?m)^(\d+) accuracy: ([0-9.]+) %, SEM: ([0-9.]+) %\s*$"
)
AGENT_RE = re.compile(r"agent (\d+)")


def parse_args():
    parser = argparse.ArgumentParser(
        description="Evaluate all MAPR result files and export an Excel report."
    )
    parser.add_argument(
        "--result-dir",
        type=Path,
        default=PROJECT_DIR / "result_qwen3_8b",
    )
    parser.add_argument(
        "--time-flag",
        help="Backward-compatible single result suffix; overrides --time-flags.",
    )
    parser.add_argument(
        "--time-flags",
        default="0713,0716,0718",
        help="Comma-separated independent run suffixes (default: 0713,0716,0718).",
    )
    parser.add_argument(
        "--interval",
        choices=["std", "ci95"],
        default="std",
        help="Radius after ±: sample standard deviation (std) or 95%% CI half-width.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        help="Output .xlsx path. Defaults inside result-dir.",
    )
    parser.add_argument(
        "--log-dir",
        type=Path,
        help="Directory for raw eval.py logs. Defaults to result-dir/eval_logs.",
    )
    return parser.parse_args()


def detect_time_flag(result_dir):
    flags = []
    for path in result_dir.glob("*/*.json"):
        match = re.search(r"_(\d{4})\.json$", path.name)
        if match:
            flags.append(match.group(1))
    if not flags:
        raise RuntimeError(f"No result JSON files found under {result_dir}")
    counts = Counter(flags)
    flag, count = counts.most_common(1)[0]
    if len(counts) > 1:
        print(f"Detected multiple time flags {dict(counts)}; using {flag} ({count} files).")
    return flag


def expected_file(result_dir, task, method, time_flag):
    file_method = FILE_METHOD.get(method, method)
    return result_dir / task / f"{task}_{file_method}_{TASK_SIZES[task]}_{time_flag}.json"


def validate_inputs(result_dir, time_flag):
    errors = []
    checked = set()
    task_exclusions = {}
    for task, expected_size in TASK_SIZES.items():
        exclusion_sets = []
        for method, _ in METHODS:
            path = expected_file(result_dir, task, method, time_flag)
            if path in checked:
                continue
            checked.add(path)
            if not path.exists():
                errors.append(f"Missing: {path}")
                continue
            try:
                with path.open(encoding="utf-8") as handle:
                    rows = json.load(handle)
                actual_size = len(rows)
            except Exception as exc:
                errors.append(f"Unreadable JSON: {path}: {exc}")
                continue
            if actual_size != expected_size:
                errors.append(
                    f"Incomplete: {path} has {actual_size}, expected {expected_size}"
                )
            exclusions = {
                (dataset or task, index)
                for dataset, index in sample_filter.excluded_keys(rows)
            }
            exclusion_sets.append((path, exclusions))
        if exclusion_sets:
            expected_exclusions = exclusion_sets[0][1]
            for path, exclusions in exclusion_sets[1:]:
                if exclusions != expected_exclusions:
                    errors.append(
                        f"Inconsistent exclusions for {task}: {path} has "
                        f"{sorted(exclusions)}, expected {sorted(expected_exclusions)}"
                    )
            task_exclusions[task] = expected_exclusions
    if errors:
        raise RuntimeError("Input validation failed:\n" + "\n".join(errors))
    return task_exclusions


def parse_agent_metrics(output, sample_size):
    details = []
    current_agent = None
    for line in output.splitlines():
        agent_match = AGENT_RE.search(line)
        if "mean multi-agent" in line:
            current_agent = None
        elif agent_match and "---" in line:
            current_agent = int(agent_match.group(1)) + 1
        metric_match = METRIC_RE.match(line.strip())
        if metric_match and current_agent is not None:
            count, accuracy, sem = metric_match.groups()
            if int(count) == sample_size:
                details.append(
                    {
                        "agent": current_agent,
                        "accuracy": float(accuracy) / 100,
                        "sem": float(sem) / 100,
                    }
                )
    return details


def run_evaluations(result_dir, time_flag, log_dir, task_exclusions):
    log_dir.mkdir(parents=True, exist_ok=True)
    results = []
    agent_results = []
    total = len(TASK_SIZES) * len(METHODS)
    current = 0

    for task, sample_size in TASK_SIZES.items():
        effective_size = sample_size - len(task_exclusions.get(task, set()))
        for method, label in METHODS:
            current += 1
            print(f"[{current:02d}/{total}] {task} / {method}")
            command = [
                sys.executable,
                str(PROJECT_DIR / "eval.py"),
                "--eval_dir",
                str(result_dir),
                "--task",
                task,
                "--method",
                method,
                "--time_flag",
                time_flag,
            ]
            for dataset, index in sorted(task_exclusions.get(task, set())):
                command.extend(["--exclude-sample", f"{dataset}:{index}"])
            completed = subprocess.run(
                command,
                cwd=PROJECT_DIR,
                text=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
            )
            log_path = log_dir / f"{task}_{method}_{time_flag}.log"
            log_path.write_text(completed.stdout, encoding="utf-8")
            if completed.returncode != 0:
                raise RuntimeError(
                    f"eval.py failed for {task}/{method}; see {log_path}"
                )

            metrics = METRIC_RE.findall(completed.stdout)
            if not metrics:
                raise RuntimeError(
                    f"Could not parse metrics for {task}/{method}; see {log_path}"
                )
            count, accuracy, sem = metrics[-1]
            if int(count) != effective_size:
                raise RuntimeError(
                    f"Unexpected final sample count for {task}/{method}: {count}"
                )

            source_file = expected_file(result_dir, task, method, time_flag)
            results.append(
                {
                    "task": task,
                    "time_flag": time_flag,
                    "method": method,
                    "label": label,
                    "sample_size": effective_size,
                    "accuracy": float(accuracy) / 100,
                    "sem": float(sem) / 100,
                    "source_file": str(source_file),
                    "log_file": str(log_path),
                    "command": " ".join(command),
                }
            )

            if method in {"single_agent", "self_correction"}:
                for detail in parse_agent_metrics(completed.stdout, effective_size):
                    agent_results.append(
                        {
                            "task": task,
                            "time_flag": time_flag,
                            "method": method,
                            "label": label,
                            "sample_size": effective_size,
                            **detail,
                        }
                    )
    return results, agent_results


def aggregate_results(results, interval):
    grouped = {}
    for row in results:
        grouped.setdefault((row["method"], row["task"]), []).append(row)
    aggregated = []
    for (method, task), rows in grouped.items():
        accuracies = [row["accuracy"] for row in rows]
        mean = statistics.mean(accuracies)
        std = statistics.stdev(accuracies) if len(accuracies) > 1 else 0.0
        radius = std if interval == "std" else 4.3026527299 * std / math.sqrt(len(rows))
        aggregated.append(
            {
                "task": task,
                "method": method,
                "label": rows[0]["label"],
                "sample_size": rows[0]["sample_size"],
                "runs": len(rows),
                "mean_accuracy": mean,
                "radius": radius,
                "run_accuracies": accuracies,
            }
        )
    return aggregated


def style_table(ws, header_row, min_col, max_col, max_row):
    navy = "17365D"
    pale_blue = "DCE6F1"
    thin_gray = Side(style="thin", color="D9E1F2")
    for cell in ws[header_row][min_col - 1 : max_col]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(
        min_row=header_row + 1,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    ):
        for cell in row:
            cell.border = Border(bottom=thin_gray)
    if max_row > header_row:
        for cell in ws[max_row]:
            if min_col <= cell.column <= max_col:
                cell.fill = PatternFill("solid", fgColor=pale_blue)


def autosize(ws, caps=None):
    caps = caps or {}
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        width = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        ws.column_dimensions[letter].width = min(max(width + 2, 10), caps.get(letter, 45))


def build_workbook(results, aggregated, agent_results, result_dir, time_flags, interval, output_path):
    wb = Workbook()
    summary = wb.active
    summary.title = "Paper Table"
    summary.sheet_view.showGridLines = False
    tasks = list(TASK_SIZES)
    lookup = {(row["method"], row["task"]): row for row in aggregated}
    summary.append(["", "GSM8K", "SVAMP", "AQuA", "MultiArith", "AddSub", "SingleEq", "ARC-c", "StrategyQA", "Colored Objects", "Penguins"])
    for label, method in PAPER_ROWS:
        values = [
            f'{lookup[(method, task)]["mean_accuracy"] * 100:.2f} ± '
            f'{lookup[(method, task)]["radius"] * 100:.2f}'
            for task in tasks
        ]
        summary.append([label, *values])
    summary.freeze_panes = "B2"
    summary.column_dimensions["A"].width = 28
    for column in range(2, 12):
        summary.column_dimensions[get_column_letter(column)].width = 12
    summary.column_dimensions["J"].width = 17
    for row in summary.iter_rows(min_row=2, max_row=summary.max_row, min_col=2, max_col=11):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")
    for row in summary.iter_rows(min_row=1, max_row=summary.max_row, min_col=1, max_col=11):
        for cell in row:
            cell.font = Font(name="Times New Roman", size=12)
            cell.alignment = Alignment(vertical="center", horizontal="left" if cell.column == 1 else "center")
    for cell in summary[1]:
        cell.font = Font(name="Times New Roman", size=12, bold=True)
    separator = Side(style="thin", color="000000")
    for cell in summary[5]:
        cell.border = Border(bottom=separator)
    for row_number in range(1, 9):
        summary.row_dimensions[row_number].height = 22

    detail = wb.create_sheet("Evaluation Results")
    detail.sheet_view.showGridLines = False
    headers = [
        "Dataset",
        "Run",
        "Method",
        "Method Label",
        "Samples",
        "Accuracy",
        "SEM",
        "Source JSON",
        "Eval Log",
        "Command",
    ]
    detail.append(headers)
    for row in results:
        detail.append(
            [
                row["task"],
                row["time_flag"],
                row["method"],
                row["label"],
                row["sample_size"],
                row["accuracy"],
                row["sem"],
                row["source_file"],
                row["log_file"],
                row["command"],
            ]
        )
    style_table(detail, 1, 1, len(headers), detail.max_row)
    detail.freeze_panes = "A2"
    detail.auto_filter.ref = detail.dimensions
    for row in detail.iter_rows(min_row=2, min_col=6, max_col=7):
        for cell in row:
            cell.number_format = "0.00%"
    autosize(detail, {"D": 28, "H": 60, "I": 55, "J": 80})

    aggregate = wb.create_sheet("Aggregate Results")
    aggregate_headers = ["Dataset", "Method", "Method Label", "Samples", "Runs", "Mean Accuracy", "Radius", "Interval"]
    aggregate.append(aggregate_headers)
    for row in aggregated:
        aggregate.append([row["task"], row["method"], row["label"], row["sample_size"], row["runs"], row["mean_accuracy"], row["radius"], interval])
    style_table(aggregate, 1, 1, len(aggregate_headers), aggregate.max_row)
    for row in aggregate.iter_rows(min_row=2, min_col=6, max_col=7):
        for cell in row:
            cell.number_format = "0.00%"
    aggregate.freeze_panes = "A2"
    aggregate.auto_filter.ref = aggregate.dimensions
    autosize(aggregate, {"C": 28})

    agents = wb.create_sheet("Agent Details")
    agents.sheet_view.showGridLines = False
    agent_headers = ["Dataset", "Run", "Method", "Method Label", "Samples", "Agent", "Accuracy", "SEM"]
    agents.append(agent_headers)
    for row in agent_results:
        agents.append(
            [
                row["task"],
                row["time_flag"],
                row["method"],
                row["label"],
                row["sample_size"],
                row["agent"],
                row["accuracy"],
                row["sem"],
            ]
        )
    style_table(agents, 1, 1, len(agent_headers), agents.max_row)
    agents.freeze_panes = "A2"
    agents.auto_filter.ref = agents.dimensions
    for row in agents.iter_rows(min_row=2, min_col=7, max_col=8):
        for cell in row:
            cell.number_format = "0.00%"
    autosize(agents, {"C": 28})

    readme = wb.create_sheet("README")
    readme.sheet_view.showGridLines = False
    readme.append(["Field", "Value"])
    readme_rows = [
        ("Generated", datetime.now().astimezone().isoformat(timespec="seconds")),
        ("Result directory", str(result_dir)),
        ("Time flags", ", ".join(time_flags)),
        ("Aggregation", f"Mean across {len(time_flags)} independent runs; ± radius uses {interval}."),
        ("Evaluator", str(PROJECT_DIR / "eval.py")),
        ("Evaluation count", len(results)),
        ("Source JSON count", len({row["source_file"] for row in results})),
        ("Majority", "Reuses each dataset's single_agent JSON; no separate generation file."),
        ("Accuracy", "Fraction of correctly evaluated examples."),
        ("SEM", "Population standard deviation divided by sqrt(number of examples), as implemented in eval.py."),
        ("Paper Table", "Dataset header row plus method rows. Each cell is mean ± radius on the 0-100 scale, without a percent sign."),
        ("Ours (w/o solution)", "Revision receives peer feedback but Stage-2 peer-solution dialogue is removed."),
    ]
    for item in readme_rows:
        readme.append(list(item))
    style_table(readme, 1, 1, 2, readme.max_row)
    readme.freeze_panes = "A2"
    readme.column_dimensions["A"].width = 24
    readme.column_dimensions["B"].width = 100
    for row in readme.iter_rows(min_row=2, max_col=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def verify_workbook(output_path, expected_results):
    wb = load_workbook(output_path, data_only=False, read_only=True)
    required = {"Paper Table", "Evaluation Results", "Aggregate Results", "Agent Details", "README"}
    if set(wb.sheetnames) != required:
        raise RuntimeError(f"Unexpected workbook sheets: {wb.sheetnames}")
    if wb["Evaluation Results"].max_row != expected_results + 1:
        raise RuntimeError("Excel detail row count does not match evaluation count")
    paper = wb["Paper Table"]
    if paper.max_row != 8 or paper.max_column != 11:
        raise RuntimeError("Paper table shape verification failed")
    expected_headers = ["GSM8K", "SVAMP", "AQuA", "MultiArith", "AddSub", "SingleEq", "ARC-c", "StrategyQA", "Colored Objects", "Penguins"]
    actual_headers = [paper.cell(1, column).value for column in range(2, 12)]
    if actual_headers != expected_headers:
        raise RuntimeError("Paper table dataset header order verification failed")
    if paper["A2"].value != "Zero-shot CoT" or paper["A8"].value != "Ours (w/o solution)":
        raise RuntimeError("Paper table method labels verification failed")
    if " ± " not in str(paper["B2"].value) or " ± " not in str(paper["B8"].value):
        raise RuntimeError("Paper table cells are not formatted as mean ± radius")
    wb.close()


def main():
    args = parse_args()
    result_dir = args.result_dir.resolve()
    time_flags = [args.time_flag] if args.time_flag else [
        value.strip() for value in args.time_flags.split(",") if value.strip()
    ]
    if not time_flags:
        raise RuntimeError("No time flags selected")
    if len(set(time_flags)) != len(time_flags):
        raise RuntimeError(f"Duplicate time flags: {time_flags}")
    flag_label = "-".join(time_flags)
    output_path = (
        args.output.resolve()
        if args.output
        else result_dir / f"evaluation_{result_dir.name}_{flag_label}.xlsx"
    )
    log_dir = (
        args.log_dir.resolve() if args.log_dir else result_dir / "eval_logs" / flag_label
    )

    exclusions_by_flag = {
        flag: validate_inputs(result_dir, flag) for flag in time_flags
    }
    shared_exclusions = {
        task: set().union(*(exclusions_by_flag[flag].get(task, set()) for flag in time_flags))
        for task in TASK_SIZES
    }
    results, agent_results = [], []
    for flag in time_flags:
        run_results, run_agents = run_evaluations(
            result_dir, flag, log_dir, shared_exclusions
        )
        results.extend(run_results)
        agent_results.extend(run_agents)
    aggregated = aggregate_results(results, args.interval)
    build_workbook(results, aggregated, agent_results, result_dir, time_flags, args.interval, output_path)
    verify_workbook(output_path, len(results))
    print(f"Excel report: {output_path}")


if __name__ == "__main__":
    main()
